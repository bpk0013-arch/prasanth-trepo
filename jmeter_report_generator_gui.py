#!/usr/bin/env python3
"""
JMeter Report Generator - GUI Version
User-friendly interface for generating JMeter performance test reports
"""

import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import threading
from datetime import datetime

# Import the report generation logic
sys.path.insert(0, 'old-data-work')
from jmeter_report_generator import JMeterResultParser, generate_pdf_report, format_duration

class JMeterReportGeneratorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("JMeter Report Generator - by Prasanth Kumar")
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        
        # Configure style
        style = ttk.Style()
        style.theme_use('clam')
        
        # Variables
        self.input_file = tk.StringVar()
        self.output_dir = tk.StringVar(value=os.path.join(os.getcwd(), "test_reports"))
        self.project_name = tk.StringVar(value="BIE Performance Test")
        self.environment = tk.StringVar(value="Test Environment")
        self.tester_name = tk.StringVar(value="Prasanth Kumar Birupogu")
        self.generate_pdf_var = tk.BooleanVar(value=True)
        self.generate_excel_var = tk.BooleanVar(value=True)
        
        self.create_widgets()
        
    def create_widgets(self):
        # Main container
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="JMeter Report Generator", 
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Input File Section
        row = 1
        ttk.Label(main_frame, text="Input JTL/CSV File:", 
                 font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=5)
        
        ttk.Entry(main_frame, textvariable=self.input_file, 
                 width=60).grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5)
        
        ttk.Button(main_frame, text="Browse...", 
                  command=self.browse_input).grid(row=row, column=2, padx=5)
        
        # Output Directory Section
        row += 1
        ttk.Label(main_frame, text="Output Directory:", 
                 font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=5)
        
        ttk.Entry(main_frame, textvariable=self.output_dir, 
                 width=60).grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5)
        
        ttk.Button(main_frame, text="Browse...", 
                  command=self.browse_output).grid(row=row, column=2, padx=5)
        
        # Separator
        row += 1
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, 
                                                            columnspan=3, sticky=(tk.W, tk.E), pady=15)
        
        # Report Details Section
        row += 1
        ttk.Label(main_frame, text="Report Details", 
                 font=('Arial', 12, 'bold')).grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(0, 10))
        
        # Project Name
        row += 1
        ttk.Label(main_frame, text="Project Name:").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.project_name, 
                 width=60).grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5, columnspan=2)
        
        # Environment
        row += 1
        ttk.Label(main_frame, text="Environment:").grid(row=row, column=0, sticky=tk.W, pady=5)
        env_combo = ttk.Combobox(main_frame, textvariable=self.environment, 
                                values=["Test Environment", "Development", "QA", "Staging", "Production"],
                                width=57)
        env_combo.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5, columnspan=2)
        
        # Tester Name
        row += 1
        ttk.Label(main_frame, text="Tester Name:").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.tester_name, 
                 width=60).grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5, columnspan=2)
        
        # Report Format Options
        row += 1
        ttk.Label(main_frame, text="Report Formats:", 
                 font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=(10, 5))
        
        format_frame = ttk.Frame(main_frame)
        format_frame.grid(row=row, column=1, sticky=tk.W, padx=5, columnspan=2)
        
        ttk.Checkbutton(format_frame, text="Generate PDF Report", 
                       variable=self.generate_pdf_var).pack(side=tk.LEFT, padx=(0, 20))
        ttk.Checkbutton(format_frame, text="Generate Excel Report", 
                       variable=self.generate_excel_var).pack(side=tk.LEFT)
        
        # Separator
        row += 1
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, 
                                                            columnspan=3, sticky=(tk.W, tk.E), pady=15)
        
        # Buttons
        row += 1
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=row, column=0, columnspan=3, pady=10)
        
        self.generate_btn = ttk.Button(button_frame, text="Generate Reports", 
                                       command=self.generate_reports,
                                       style='Accent.TButton')
        self.generate_btn.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="Clear Log", 
                  command=self.clear_log).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="Open Output Folder", 
                  command=self.open_output_folder).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="Exit", 
                  command=self.root.quit).pack(side=tk.LEFT, padx=5)
        
        # Log Section
        row += 1
        ttk.Label(main_frame, text="Log Output:", 
                 font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=(10, 5))
        
        row += 1
        self.log_text = ScrolledText(main_frame, height=15, width=100, 
                                     wrap=tk.WORD, font=('Consolas', 9))
        self.log_text.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # Configure row weight for log expansion
        main_frame.rowconfigure(row, weight=1)
        
        # Status Bar
        row += 1
        self.status_label = ttk.Label(main_frame, text="Ready", relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # Initial log message
        self.log("Welcome to JMeter Report Generator!")
        self.log("Select a JTL/CSV file and click 'Generate Reports' to begin.")
        self.log("-" * 80)
        
    def browse_input(self):
        filename = filedialog.askopenfilename(
            title="Select JMeter Result File",
            filetypes=[
                ("JMeter Files", "*.jtl *.csv"),
                ("All Files", "*.*")
            ]
        )
        if filename:
            self.input_file.set(filename)
            self.log(f"Selected input file: {filename}")
            
    def browse_output(self):
        directory = filedialog.askdirectory(title="Select Output Directory")
        if directory:
            self.output_dir.set(directory)
            self.log(f"Output directory set to: {directory}")
            
    def log(self, message):
        """Add message to log with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def clear_log(self):
        """Clear the log text"""
        self.log_text.delete(1.0, tk.END)
        self.log("Log cleared.")
        
    def open_output_folder(self):
        """Open the output folder in file explorer"""
        output_dir = self.output_dir.get()
        if os.path.exists(output_dir):
            if sys.platform == 'win32':
                os.startfile(output_dir)
            elif sys.platform == 'darwin':
                os.system(f'open "{output_dir}"')
            else:
                os.system(f'xdg-open "{output_dir}"')
            self.log(f"Opened output folder: {output_dir}")
        else:
            messagebox.showwarning("Folder Not Found", 
                                 f"Output folder does not exist:\n{output_dir}")
            
    def update_status(self, message):
        """Update status bar"""
        self.status_label.config(text=message)
        self.root.update_idletasks()
        
    def validate_inputs(self):
        """Validate user inputs"""
        if not self.input_file.get():
            messagebox.showerror("Error", "Please select an input JTL/CSV file.")
            return False
            
        if not os.path.exists(self.input_file.get()):
            messagebox.showerror("Error", "Input file does not exist.")
            return False
            
        if not self.output_dir.get():
            messagebox.showerror("Error", "Please select an output directory.")
            return False
            
        if not self.generate_pdf_var.get() and not self.generate_excel_var.get():
            messagebox.showerror("Error", "Please select at least one report format.")
            return False
            
        return True
        
    def generate_reports(self):
        """Generate reports in a separate thread"""
        if not self.validate_inputs():
            return
            
        # Disable generate button
        self.generate_btn.config(state='disabled')
        self.update_status("Generating reports...")
        
        # Run in separate thread to keep UI responsive
        thread = threading.Thread(target=self._generate_reports_thread, daemon=True)
        thread.start()
        
    def _generate_reports_thread(self):
        """Thread function for report generation"""
        try:
            input_file = self.input_file.get()
            output_dir = self.output_dir.get()
            
            self.log("=" * 80)
            self.log("📊 JMeter Report Generator")
            self.log("=" * 80)
            self.log(f"Input file: {input_file}")
            self.log(f"Project: {self.project_name.get()}")
            self.log(f"Environment: {self.environment.get()}")
            self.log(f"Tester: {self.tester_name.get()}")
            self.log("")
            
            # Create output directory if it doesn't exist
            os.makedirs(output_dir, exist_ok=True)
            
            # Parse JMeter results
            self.log("📖 Parsing JMeter results...")
            parser = JMeterResultParser(input_file)
            transactions, overall = parser.parse()
            
            self.log(f"✅ Parsed {overall['totalSamples']} samples from {len(transactions)} transactions")
            self.log("")
            
            # Generate timestamp for filenames
            timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
            base_name = f"Performance_Test_Report_{timestamp}"
            
            # Generate PDF
            if self.generate_pdf_var.get():
                pdf_path = os.path.join(output_dir, f"{base_name}.pdf")
                self.log(f"📄 Generating PDF report: {pdf_path}")
                
                generate_pdf_report(
                    transactions, overall, pdf_path,
                    self.project_name.get(),
                    self.environment.get(),
                    self.tester_name.get()
                )
                
                self.log("✅ PDF report generated successfully!")
                self.log("")
            
            # Generate Excel
            if self.generate_excel_var.get():
                excel_path = os.path.join(output_dir, f"{base_name}.xlsx")
                self.log(f"📊 Generating Excel report: {excel_path}")
                
                # Import Excel generation
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Summary"
                
                # Add summary data
                ws['A1'] = 'Performance Test Report'
                ws['A1'].font = Font(size=16, bold=True)
                
                row = 3
                ws[f'A{row}'] = 'Project Name:'
                ws[f'B{row}'] = self.project_name.get()
                row += 1
                ws[f'A{row}'] = 'Environment:'
                ws[f'B{row}'] = self.environment.get()
                row += 1
                ws[f'A{row}'] = 'Tester:'
                ws[f'B{row}'] = self.tester_name.get()
                row += 1
                ws[f'A{row}'] = 'Test Start:'
                ws[f'B{row}'] = overall['testStart'].strftime('%Y-%m-%d %H:%M:%S')
                row += 1
                ws[f'A{row}'] = 'Test End:'
                ws[f'B{row}'] = overall['testEnd'].strftime('%Y-%m-%d %H:%M:%S')
                row += 1
                ws[f'A{row}'] = 'Duration:'
                ws[f'B{row}'] = format_duration(overall['duration'])
                
                row += 2
                ws[f'A{row}'] = 'Total Samples:'
                ws[f'B{row}'] = overall['totalSamples']
                row += 1
                ws[f'A{row}'] = 'Total Errors:'
                ws[f'B{row}'] = overall['totalErrors']
                row += 1
                ws[f'A{row}'] = 'Error Rate:'
                ws[f'B{row}'] = f"{overall['errorRate']:.2f}%"
                row += 1
                ws[f'A{row}'] = 'Avg Response Time:'
                ws[f'B{row}'] = f"{overall['avgResponseTime']:.2f} ms"
                
                # Add detailed results sheet
                ws2 = wb.create_sheet("Detailed Results")
                headers = ['Transaction', 'Samples', 'Errors', 'Error%', 'Avg(ms)', 
                          'Min(ms)', 'Max(ms)', '90%(ms)', '95%(ms)', '99%(ms)']
                ws2.append(headers)
                
                for tx in transactions:
                    ws2.append([
                        tx['name'],
                        tx['samples'],
                        tx['errors'],
                        f"{tx['errorPct']:.2f}",
                        f"{tx['avgTime']:.2f}",
                        tx['minTime'],
                        tx['maxTime'],
                        tx['p90'],
                        tx['p95'],
                        tx['p99']
                    ])
                
                # Style headers
                for cell in ws2[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                
                wb.save(excel_path)
                self.log("✅ Excel report generated successfully!")
                self.log("")
            
            self.log("🎉 Report generation complete!")
            self.log(f"📁 Reports saved to: {output_dir}")
            self.log("=" * 80)
            
            # Show success message
            self.root.after(0, lambda: messagebox.showinfo(
                "Success", 
                f"Reports generated successfully!\n\nLocation: {output_dir}"
            ))
            
            self.root.after(0, lambda: self.update_status("Reports generated successfully!"))
            
        except Exception as e:
            error_msg = f"Error generating reports: {str(e)}"
            self.log(f"❌ {error_msg}")
            self.log("")
            
            import traceback
            self.log("Traceback:")
            self.log(traceback.format_exc())
            
            self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
            self.root.after(0, lambda: self.update_status("Error occurred"))
            
        finally:
            # Re-enable generate button
            self.root.after(0, lambda: self.generate_btn.config(state='normal'))


def main():
    root = tk.Tk()
    app = JMeterReportGeneratorGUI(root)
    
    # Center window on screen
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()


if __name__ == '__main__':
    main()
