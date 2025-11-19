#!/usr/bin/env python3
"""
JMeter Report Generator - Standalone GUI Version
All-in-one file with no external dependencies (except libraries)
"""

import os
import sys
import csv
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import threading
from datetime import datetime
from collections import defaultdict
import math

# PDF Generation imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER

# Excel Generation imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================================
# JMeter Result Parser (embedded)
# ============================================================================

class JMeterResultParser:
    """Parse JMeter JTL/CSV result files"""
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.transactions = defaultdict(lambda: {
            'samples': [],
            'errors': 0,
            'start_time': None,
            'end_time': None
        })
        
    def parse(self):
        """Parse JTL/CSV file and extract metrics"""
        with open(self.file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            
            for row in reader:
                label = row.get('label', row.get('Label', 'Unknown'))
                elapsed = int(row.get('elapsed', row.get('Elapsed', 0)))
                success = row.get('success', row.get('Success', 'true')).lower() == 'true'
                timestamp = int(row.get('timeStamp', row.get('TimeStamp', 0)))
                
                tx = self.transactions[label]
                tx['samples'].append(elapsed)
                
                if not success:
                    tx['errors'] += 1
                
                if tx['start_time'] is None or timestamp < tx['start_time']:
                    tx['start_time'] = timestamp
                
                if tx['end_time'] is None or timestamp > tx['end_time']:
                    tx['end_time'] = timestamp + elapsed
        
        return self.calculate_metrics()
    
    def calculate_metrics(self):
        """Calculate statistics for each transaction"""
        results = []
        total_samples = 0
        total_errors = 0
        total_response_time = 0
        test_start = None
        test_end = None
        
        for label, data in self.transactions.items():
            samples = data['samples']
            if not samples:
                continue
            
            samples_sorted = sorted(samples)
            sample_count = len(samples)
            error_count = data['errors']
            
            # Calculate percentiles
            p90_idx = int(0.90 * sample_count)
            p95_idx = int(0.95 * sample_count)
            p99_idx = int(0.99 * sample_count)
            
            # Calculate throughput
            duration_ms = data['end_time'] - data['start_time']
            throughput = (sample_count * 1000.0 / duration_ms) if duration_ms > 0 else 0
            
            metrics = {
                'name': label,
                'samples': sample_count,
                'errors': error_count,
                'errorPct': (error_count * 100.0 / sample_count) if sample_count > 0 else 0,
                'avgTime': sum(samples) / sample_count,
                'minTime': min(samples),
                'maxTime': max(samples),
                'p90': samples_sorted[p90_idx] if p90_idx < sample_count else samples_sorted[-1],
                'p95': samples_sorted[p95_idx] if p95_idx < sample_count else samples_sorted[-1],
                'p99': samples_sorted[p99_idx] if p99_idx < sample_count else samples_sorted[-1],
                'throughput': throughput
            }
            
            results.append(metrics)
            
            total_samples += sample_count
            total_errors += error_count
            total_response_time += sum(samples)
            
            if test_start is None or data['start_time'] < test_start:
                test_start = data['start_time']
            if test_end is None or data['end_time'] > test_end:
                test_end = data['end_time']
        
        # Overall metrics
        overall = {
            'totalSamples': total_samples,
            'totalErrors': total_errors,
            'errorRate': (total_errors * 100.0 / total_samples) if total_samples > 0 else 0,
            'avgResponseTime': (total_response_time / total_samples) if total_samples > 0 else 0,
            'testStart': datetime.fromtimestamp(test_start / 1000) if test_start else datetime.now(),
            'testEnd': datetime.fromtimestamp(test_end / 1000) if test_end else datetime.now(),
            'duration': test_end - test_start if test_start and test_end else 0
        }
        
        return results, overall

# ============================================================================
# Report Generation Functions (embedded)
# ============================================================================

def format_duration(millis):
    """Format duration from milliseconds to HH:MM:SS"""
    seconds = millis // 1000
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"

def generate_pdf_report(transactions, overall, output_path, project_name, environment, tester_name):
    """Generate PDF report"""
    doc = SimpleDocTemplate(output_path, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    title = Paragraph("Performance Test Report", title_style)
    story.append(title)
    story.append(Spacer(1, 0.2*inch))
    
    # Test Information
    info_data = [
        ['Project Name:', project_name],
        ['Test Environment:', environment],
        ['Tester Name:', tester_name],
        ['Test Start Time:', overall['testStart'].strftime('%Y-%m-%d %H:%M:%S')],
        ['Test End Time:', overall['testEnd'].strftime('%Y-%m-%d %H:%M:%S')],
        ['Test Duration:', format_duration(overall['duration'])],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2C3E50')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Executive Summary
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#34495E'),
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    summary_heading = Paragraph("Executive Summary", heading_style)
    story.append(summary_heading)
    
    status = 'PASSED' if overall['errorRate'] < 1.0 else 'FAILED'
    status_color = colors.green if status == 'PASSED' else colors.red
    
    summary_data = [
        ['Total Samples:', str(overall['totalSamples'])],
        ['Total Errors:', str(overall['totalErrors'])],
        ['Error Rate:', f"{overall['errorRate']:.2f}%"],
        ['Average Response Time:', f"{overall['avgResponseTime']:.2f} ms"],
        ['Test Status:', status],
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
    summary_table.setStyle(TableStyle([
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('FONT', (1, 0), (1, 3), 'Helvetica', 10),
        ('FONT', (1, 4), (1, 4), 'Helvetica-Bold', 10),
        ('TEXTCOLOR', (0, 0), (-1, 3), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (1, 4), (1, 4), status_color),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#ECF0F1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Detailed Results
    results_heading = Paragraph("Detailed Transaction Results", heading_style)
    story.append(results_heading)
    
    results_data = [
        ['Transaction', 'Samples', 'Errors', 'Error%', 'Avg(ms)', 
         'Min(ms)', 'Max(ms)', '90%(ms)', '95%(ms)', '99%(ms)']
    ]
    
    for tx in transactions:
        results_data.append([
            tx['name'],
            str(tx['samples']),
            str(tx['errors']),
            f"{tx['errorPct']:.2f}",
            f"{tx['avgTime']:.2f}",
            str(tx['minTime']),
            str(tx['maxTime']),
            str(tx['p90']),
            str(tx['p95']),
            str(tx['p99'])
        ])
    
    results_table = Table(results_data, colWidths=[1.5*inch] + [0.6*inch]*9)
    results_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 8),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2C3E50')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    story.append(results_table)
    
    doc.build(story)

# ============================================================================
# GUI Application
# ============================================================================

class JMeterReportGeneratorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("JMeter Report Generator - by Prasanth Kumar")
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        
        # Configure style
        style = ttk.Style()
        style.theme_use('clam')
        
        # Variables
        self.input_file = tk.StringVar()
        self.output_dir = tk.StringVar(value=os.path.join(os.getcwd(), "test_reports"))
        self.project_name = tk.StringVar(value="BIE Performance Test")
        self.environment = tk.StringVar(value="Test Environment")
        self.tester_name = tk.StringVar(value="Prasanth Kumar Birupogu")
        self.generate_pdf_var = tk.BooleanVar(value=True)
        self.generate_excel_var = tk.BooleanVar(value=True)
        
        self.create_widgets()
        
    def create_widgets(self):
        # Main container
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="JMeter Report Generator", 
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Input File Section
        row = 1
        ttk.Label(main_frame, text="Input JTL/CSV File:", 
                 font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=5)
        
        ttk.Entry(main_frame, textvariable=self.input_file, 
                 width=60).grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5)
        
        ttk.Button(main_frame, text="Browse...", 
                  command=self.browse_input).grid(row=row, column=2, padx=5)
        
        # Output Directory Section
        row += 1
        ttk.Label(main_frame, text="Output Directory:", 
                 font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=5)
        
        ttk.Entry(main_frame, textvariable=self.output_dir, 
                 width=60).grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5)
        
        ttk.Button(main_frame, text="Browse...", 
                  command=self.browse_output).grid(row=row, column=2, padx=5)
        
        # Separator
        row += 1
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, 
                                                            columnspan=3, sticky=(tk.W, tk.E), pady=15)
        
        # Report Details Section
        row += 1
        ttk.Label(main_frame, text="Report Details", 
                 font=('Arial', 12, 'bold')).grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(0, 10))
        
        # Project Name
        row += 1
        ttk.Label(main_frame, text="Project Name:").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.project_name, 
                 width=60).grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5, columnspan=2)
        
        # Environment
        row += 1
        ttk.Label(main_frame, text="Environment:").grid(row=row, column=0, sticky=tk.W, pady=5)
        env_combo = ttk.Combobox(main_frame, textvariable=self.environment, 
                                values=["Test Environment", "Development", "QA", "Staging", "Production"],
                                width=57)
        env_combo.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5, columnspan=2)
        
        # Tester Name
        row += 1
        ttk.Label(main_frame, text="Tester Name:").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.tester_name, 
                 width=60).grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5, columnspan=2)
        
        # Report Format Options
        row += 1
        ttk.Label(main_frame, text="Report Formats:", 
                 font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=(10, 5))
        
        format_frame = ttk.Frame(main_frame)
        format_frame.grid(row=row, column=1, sticky=tk.W, padx=5, columnspan=2)
        
        ttk.Checkbutton(format_frame, text="Generate PDF Report", 
                       variable=self.generate_pdf_var).pack(side=tk.LEFT, padx=(0, 20))
        ttk.Checkbutton(format_frame, text="Generate Excel Report", 
                       variable=self.generate_excel_var).pack(side=tk.LEFT)
        
        # Separator
        row += 1
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, 
                                                            columnspan=3, sticky=(tk.W, tk.E), pady=15)
        
        # Buttons
        row += 1
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=row, column=0, columnspan=3, pady=10)
        
        self.generate_btn = ttk.Button(button_frame, text="Generate Reports", 
                                       command=self.generate_reports,
                                       style='Accent.TButton')
        self.generate_btn.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="Clear Log", 
                  command=self.clear_log).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="Open Output Folder", 
                  command=self.open_output_folder).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="Exit", 
                  command=self.root.quit).pack(side=tk.LEFT, padx=5)
        
        # Log Section
        row += 1
        ttk.Label(main_frame, text="Log Output:", 
                 font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=(10, 5))
        
        row += 1
        self.log_text = ScrolledText(main_frame, height=15, width=100, 
                                     wrap=tk.WORD, font=('Consolas', 9))
        self.log_text.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # Configure row weight for log expansion
        main_frame.rowconfigure(row, weight=1)
        
        # Status Bar
        row += 1
        self.status_label = ttk.Label(main_frame, text="Ready", relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # Initial log message
        self.log("Welcome to JMeter Report Generator!")
        self.log("Select a JTL/CSV file and click 'Generate Reports' to begin.")
        self.log("-" * 80)
        
    def browse_input(self):
        filename = filedialog.askopenfilename(
            title="Select JMeter Result File",
            filetypes=[
                ("JMeter Files", "*.jtl *.csv"),
                ("All Files", "*.*")
            ]
        )
        if filename:
            self.input_file.set(filename)
            self.log(f"Selected input file: {filename}")
            
    def browse_output(self):
        directory = filedialog.askdirectory(title="Select Output Directory")
        if directory:
            self.output_dir.set(directory)
            self.log(f"Output directory set to: {directory}")
            
    def log(self, message):
        """Add message to log with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def clear_log(self):
        """Clear the log text"""
        self.log_text.delete(1.0, tk.END)
        self.log("Log cleared.")
        
    def open_output_folder(self):
        """Open the output folder in file explorer"""
        output_dir = self.output_dir.get()
        if os.path.exists(output_dir):
            if sys.platform == 'win32':
                os.startfile(output_dir)
            elif sys.platform == 'darwin':
                os.system(f'open "{output_dir}"')
            else:
                os.system(f'xdg-open "{output_dir}"')
            self.log(f"Opened output folder: {output_dir}")
        else:
            messagebox.showwarning("Folder Not Found", 
                                 f"Output folder does not exist:\n{output_dir}")
            
    def update_status(self, message):
        """Update status bar"""
        self.status_label.config(text=message)
        self.root.update_idletasks()
        
    def validate_inputs(self):
        """Validate user inputs"""
        if not self.input_file.get():
            messagebox.showerror("Error", "Please select an input JTL/CSV file.")
            return False
            
        if not os.path.exists(self.input_file.get()):
            messagebox.showerror("Error", "Input file does not exist.")
            return False
            
        if not self.output_dir.get():
            messagebox.showerror("Error", "Please select an output directory.")
            return False
            
        if not self.generate_pdf_var.get() and not self.generate_excel_var.get():
            messagebox.showerror("Error", "Please select at least one report format.")
            return False
            
        return True
        
    def generate_reports(self):
        """Generate reports in a separate thread"""
        if not self.validate_inputs():
            return
            
        # Disable generate button
        self.generate_btn.config(state='disabled')
        self.update_status("Generating reports...")
        
        # Run in separate thread to keep UI responsive
        thread = threading.Thread(target=self._generate_reports_thread, daemon=True)
        thread.start()
        
    def _generate_reports_thread(self):
        """Thread function for report generation"""
        try:
            input_file = self.input_file.get()
            output_dir = self.output_dir.get()
            
            self.log("=" * 80)
            self.log("📊 JMeter Report Generator")
            self.log("=" * 80)
            self.log(f"Input file: {input_file}")
            self.log(f"Project: {self.project_name.get()}")
            self.log(f"Environment: {self.environment.get()}")
            self.log(f"Tester: {self.tester_name.get()}")
            self.log("")
            
            # Create output directory if it doesn't exist
            os.makedirs(output_dir, exist_ok=True)
            
            # Parse JMeter results
            self.log("📖 Parsing JMeter results...")
            parser = JMeterResultParser(input_file)
            transactions, overall = parser.parse()
            
            self.log(f"✅ Parsed {overall['totalSamples']} samples from {len(transactions)} transactions")
            self.log("")
            
            # Generate timestamp for filenames
            timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
            base_name = f"Performance_Test_Report_{timestamp}"
            
            # Generate PDF
            if self.generate_pdf_var.get():
                pdf_path = os.path.join(output_dir, f"{base_name}.pdf")
                self.log(f"📄 Generating PDF report: {pdf_path}")
                
                generate_pdf_report(
                    transactions, overall, pdf_path,
                    self.project_name.get(),
                    self.environment.get(),
                    self.tester_name.get()
                )
                
                self.log("✅ PDF report generated successfully!")
                self.log("")
            
            # Generate Excel
            if self.generate_excel_var.get():
                excel_path = os.path.join(output_dir, f"{base_name}.xlsx")
                self.log(f"📊 Generating Excel report: {excel_path}")
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Summary"
                
                # Add summary data
                ws['A1'] = 'Performance Test Report'
                ws['A1'].font = Font(size=16, bold=True)
                
                row = 3
                ws[f'A{row}'] = 'Project Name:'
                ws[f'B{row}'] = self.project_name.get()
                row += 1
                ws[f'A{row}'] = 'Environment:'
                ws[f'B{row}'] = self.environment.get()
                row += 1
                ws[f'A{row}'] = 'Tester:'
                ws[f'B{row}'] = self.tester_name.get()
                row += 1
                ws[f'A{row}'] = 'Test Start:'
                ws[f'B{row}'] = overall['testStart'].strftime('%Y-%m-%d %H:%M:%S')
                row += 1
                ws[f'A{row}'] = 'Test End:'
                ws[f'B{row}'] = overall['testEnd'].strftime('%Y-%m-%d %H:%M:%S')
                row += 1
                ws[f'A{row}'] = 'Duration:'
                ws[f'B{row}'] = format_duration(overall['duration'])
                
                row += 2
                ws[f'A{row}'] = 'Total Samples:'
                ws[f'B{row}'] = overall['totalSamples']
                row += 1
                ws[f'A{row}'] = 'Total Errors:'
                ws[f'B{row}'] = overall['totalErrors']
                row += 1
                ws[f'A{row}'] = 'Error Rate:'
                ws[f'B{row}'] = f"{overall['errorRate']:.2f}%"
                row += 1
                ws[f'A{row}'] = 'Avg Response Time:'
                ws[f'B{row}'] = f"{overall['avgResponseTime']:.2f} ms"
                
                # Add detailed results sheet
                ws2 = wb.create_sheet("Detailed Results")
                headers = ['Transaction', 'Samples', 'Errors', 'Error%', 'Avg(ms)', 
                          'Min(ms)', 'Max(ms)', '90%(ms)', '95%(ms)', '99%(ms)']
                ws2.append(headers)
                
                for tx in transactions:
                    ws2.append([
                        tx['name'],
                        tx['samples'],
                        tx['errors'],
                        f"{tx['errorPct']:.2f}",
                        f"{tx['avgTime']:.2f}",
                        tx['minTime'],
                        tx['maxTime'],
                        tx['p90'],
                        tx['p95'],
                        tx['p99']
                    ])
                
                # Style headers
                for cell in ws2[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                
                wb.save(excel_path)
                self.log("✅ Excel report generated successfully!")
                self.log("")
            
            self.log("🎉 Report generation complete!")
            self.log(f"📁 Reports saved to: {output_dir}")
            self.log("=" * 80)
            
            # Show success message
            self.root.after(0, lambda: messagebox.showinfo(
                "Success", 
                f"Reports generated successfully!\n\nLocation: {output_dir}"
            ))
            
            self.root.after(0, lambda: self.update_status("Reports generated successfully!"))
            
        except Exception as e:
            error_msg = f"Error generating reports: {str(e)}"
            self.log(f"❌ {error_msg}")
            self.log("")
            
            import traceback
            self.log("Traceback:")
            self.log(traceback.format_exc())
            
            self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
            self.root.after(0, lambda: self.update_status("Error occurred"))
            
        finally:
            # Re-enable generate button
            self.root.after(0, lambda: self.generate_btn.config(state='normal'))


def main():
    root = tk.Tk()
    app = JMeterReportGeneratorGUI(root)
    
    # Center window on screen
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()


if __name__ == '__main__':
    main()
