#!/usr/bin/env python3
"""
JMeter Report Generator - Enhanced Professional Version
Features:
- Sky blue background theme
- Charts and graphs
- Error highlighting (red background, white bold text)
- Perfect alignment in PDF and Excel
- Detailed analysis
"""

import os
import sys
import csv
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import threading
from datetime import datetime
from collections import defaultdict
import math

# PDF Generation imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph, 
                                Spacer, PageBreak, Image as RLImage)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie

# Excel Generation imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

# Define color scheme
SKY_BLUE = colors.HexColor('#87CEEB')  # Sky blue
LIGHT_SKY_BLUE = colors.HexColor('#B0E0E6')  # Light sky blue
DARK_BLUE = colors.HexColor('#4682B4')  # Steel blue
ERROR_RED = colors.HexColor('#DC143C')  # Crimson red
WHITE = colors.white
DARK_TEXT = colors.HexColor('#2C3E50')

# Excel colors
EXCEL_SKY_BLUE = '87CEEB'
EXCEL_LIGHT_BLUE = 'B0E0E6'
EXCEL_DARK_BLUE = '4682B4'
EXCEL_ERROR_RED = 'DC143C'
EXCEL_WHITE = 'FFFFFF'

# ============================================================================
# JMeter Result Parser
# ============================================================================

class JMeterResultParser:
    """Parse JMeter JTL/CSV result files"""
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.transactions = defaultdict(lambda: {
            'samples': [],
            'errors': 0,
            'error_codes': defaultdict(int),
            'start_time': None,
            'end_time': None,
            'timestamps': [],
            'threads': []
        })
        
    def parse(self):
        """Parse JTL/CSV file and extract metrics"""
        with open(self.file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            
            for row in reader:
                label = row.get('label', row.get('Label', 'Unknown'))
                elapsed = int(row.get('elapsed', row.get('Elapsed', 0)))
                success = row.get('success', row.get('Success', 'true')).lower() == 'true'
                timestamp = int(row.get('timeStamp', row.get('TimeStamp', 0)))
                response_code = row.get('responseCode', row.get('ResponseCode', '200'))
                threads = int(row.get('allThreads', row.get('AllThreads', 1)))
                
                tx = self.transactions[label]
                tx['samples'].append(elapsed)
                tx['timestamps'].append(timestamp)
                tx['threads'].append(threads)
                
                if not success:
                    tx['errors'] += 1
                    tx['error_codes'][response_code] += 1
                
                if tx['start_time'] is None or timestamp < tx['start_time']:
                    tx['start_time'] = timestamp
                
                if tx['end_time'] is None or timestamp > tx['end_time']:
                    tx['end_time'] = timestamp + elapsed
        
        return self.calculate_metrics()
    
    def calculate_metrics(self):
        """Calculate statistics for each transaction"""
        results = []
        total_samples = 0
        total_errors = 0
        total_response_time = 0
        test_start = None
        test_end = None
        
        for label, data in self.transactions.items():
            samples = data['samples']
            if not samples:
                continue
            
            samples_sorted = sorted(samples)
            sample_count = len(samples)
            error_count = data['errors']
            
            # Calculate percentiles
            p90_idx = int(0.90 * sample_count)
            p95_idx = int(0.95 * sample_count)
            p99_idx = int(0.99 * sample_count)
            
            # Calculate throughput
            duration_ms = data['end_time'] - data['start_time']
            throughput = (sample_count * 1000.0 / duration_ms) if duration_ms > 0 else 0
            
            metrics = {
                'name': label,
                'samples': sample_count,
                'errors': error_count,
                'error_codes': dict(data['error_codes']),
                'errorPct': (error_count * 100.0 / sample_count) if sample_count > 0 else 0,
                'avgTime': sum(samples) / sample_count,
                'minTime': min(samples),
                'maxTime': max(samples),
                'p90': samples_sorted[p90_idx] if p90_idx < sample_count else samples_sorted[-1],
                'p95': samples_sorted[p95_idx] if p95_idx < sample_count else samples_sorted[-1],
                'p99': samples_sorted[p99_idx] if p99_idx < sample_count else samples_sorted[-1],
                'throughput': throughput,
                'timestamps': data['timestamps'],
                'threads': data['threads']
            }
            
            results.append(metrics)
            
            total_samples += sample_count
            total_errors += error_count
            total_response_time += sum(samples)
            
            if test_start is None or data['start_time'] < test_start:
                test_start = data['start_time']
            if test_end is None or data['end_time'] > test_end:
                test_end = data['end_time']
        
        # Overall metrics
        overall = {
            'totalSamples': total_samples,
            'totalErrors': total_errors,
            'errorRate': (total_errors * 100.0 / total_samples) if total_samples > 0 else 0,
            'avgResponseTime': (total_response_time / total_samples) if total_samples > 0 else 0,
            'testStart': datetime.fromtimestamp(test_start / 1000) if test_start else datetime.now(),
            'testEnd': datetime.fromtimestamp(test_end / 1000) if test_end else datetime.now(),
            'duration': test_end - test_start if test_start and test_end else 0
        }
        
        return results, overall

# ============================================================================
# Helper Functions
# ============================================================================

def format_duration(millis):
    """Format duration from milliseconds to HH:MM:SS"""
    seconds = millis // 1000
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"

# ============================================================================
# Chart Generation Functions
# ============================================================================

def create_response_time_chart(transactions):
    """Create response time bar chart for PDF"""
    drawing = Drawing(500, 200)
    chart = VerticalBarChart()
    chart.x = 50
    chart.y = 50
    chart.height = 125
    chart.width = 400
    
    # Data
    names = [tx['name'][:15] for tx in transactions[:10]]  # Limit to 10 for readability
    avg_times = [tx['avgTime'] for tx in transactions[:10]]
    
    chart.data = [avg_times]
    chart.categoryAxis.categoryNames = names
    chart.categoryAxis.labels.angle = 45
    chart.categoryAxis.labels.fontSize = 8
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = max(avg_times) * 1.2 if avg_times else 100
    
    chart.bars[0].fillColor = DARK_BLUE
    
    drawing.add(chart)
    return drawing

def create_error_pie_chart(transactions):
    """Create error distribution pie chart for PDF"""
    drawing = Drawing(300, 200)
    pie = Pie()
    pie.x = 100
    pie.y = 50
    pie.width = 100
    pie.height = 100
    
    # Data - show transactions with errors
    error_data = [(tx['name'], tx['errors']) for tx in transactions if tx['errors'] > 0]
    if not error_data:
        error_data = [('No Errors', 1)]
    
    pie.data = [item[1] for item in error_data]
    pie.labels = [item[0][:15] for item in error_data]
    pie.slices.strokeWidth = 0.5
    
    # Color slices
    for i in range(len(error_data)):
        pie.slices[i].fillColor = ERROR_RED if error_data[i][1] > 0 else colors.green
    
    drawing.add(pie)
    return drawing

def create_tps_chart(transactions):
    """Create Transactions Per Second chart for PDF"""
    drawing = Drawing(500, 200)
    chart = VerticalBarChart()
    chart.x = 50
    chart.y = 50
    chart.height = 125
    chart.width = 400
    
    # Data - calculate TPS for each transaction
    names = [tx['name'][:15] for tx in transactions[:10]]
    tps_values = [tx['throughput'] for tx in transactions[:10]]
    
    chart.data = [tps_values]
    chart.categoryAxis.categoryNames = names
    chart.categoryAxis.labels.angle = 45
    chart.categoryAxis.labels.fontSize = 8
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = max(tps_values) * 1.2 if tps_values else 10
    
    chart.bars[0].fillColor = colors.HexColor('#2ECC71')  # Green for TPS
    
    drawing.add(chart)
    return drawing

def create_threads_over_time_chart(transactions):
    """Create Active Threads Over Time chart for PDF"""
    drawing = Drawing(500, 200)
    chart = HorizontalLineChart()
    chart.x = 50
    chart.y = 50
    chart.height = 125
    chart.width = 400
    
    # Aggregate thread data from all transactions
    all_timestamps = []
    all_threads = []
    
    for tx in transactions:
        if tx.get('timestamps') and tx.get('threads'):
            all_timestamps.extend(tx['timestamps'])
            all_threads.extend(tx['threads'])
    
    if all_timestamps and all_threads:
        # Sort by timestamp
        sorted_data = sorted(zip(all_timestamps, all_threads))
        
        # Sample data points (take every Nth point to avoid overcrowding)
        sample_rate = max(1, len(sorted_data) // 20)
        sampled_data = sorted_data[::sample_rate]
        
        threads_data = [t[1] for t in sampled_data]
        
        chart.data = [threads_data]
        chart.lines[0].strokeColor = DARK_BLUE
        chart.lines[0].strokeWidth = 2
        
        chart.valueAxis.valueMin = 0
        chart.valueAxis.valueMax = max(threads_data) * 1.2 if threads_data else 10
    else:
        # No data available
        chart.data = [[0]]
    
    drawing.add(chart)
    return drawing

# ============================================================================
# PDF Report Generation with Enhanced Styling
# ============================================================================

def generate_pdf_report(transactions, overall, output_path, project_name, environment, tester_name, prepared_by="", company_logo=""):
    """Generate enhanced PDF report with charts and styling"""
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                           topMargin=0.75*inch, bottomMargin=0.5*inch,
                           leftMargin=0.5*inch, rightMargin=0.5*inch)
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles with sky blue theme
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=DARK_BLUE,
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=DARK_BLUE,
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    # ========== COVER PAGE ==========
    # Add company logo if provided
    if company_logo and os.path.exists(company_logo):
        try:
            logo = RLImage(company_logo, width=1.5*inch, height=0.75*inch)
            story.append(logo)
            story.append(Spacer(1, 0.2*inch))
        except:
            pass  # Skip if logo can't be loaded
    else:
        story.append(Spacer(1, 1*inch))
    
    title = Paragraph("Performance Test Report", title_style)
    story.append(title)
    story.append(Spacer(1, 0.3*inch))
    
    # Project info box with sky blue background
    project_data = [
        ['Project:', project_name],
        ['Environment:', environment],
        ['Tested By:', tester_name],
        ['Report Prepared By:', prepared_by if prepared_by else tester_name],
        ['Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    project_table = Table(project_data, colWidths=[2*inch, 4.5*inch])
    project_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), SKY_BLUE),
        ('TEXTCOLOR', (0, 0), (-1, -1), WHITE),
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 12),
        ('FONT', (1, 0), (1, -1), 'Helvetica', 12),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, WHITE),
    ]))
    
    story.append(project_table)
    story.append(PageBreak())
    
    # ========== TEST INFORMATION ==========
    info_heading = Paragraph("Test Information", heading_style)
    story.append(info_heading)
    
    info_data = [
        ['Test Start Time:', overall['testStart'].strftime('%Y-%m-%d %H:%M:%S')],
        ['Test End Time:', overall['testEnd'].strftime('%Y-%m-%d %H:%M:%S')],
        ['Test Duration:', format_duration(overall['duration'])],
        ['Total Samples:', str(overall['totalSamples'])],
        ['Total Errors:', str(overall['totalErrors'])],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4.5*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), SKY_BLUE),
        ('BACKGROUND', (1, 0), (1, -1), LIGHT_SKY_BLUE),
        ('TEXTCOLOR', (0, 0), (0, -1), WHITE),
        ('TEXTCOLOR', (1, 0), (1, -1), DARK_TEXT),
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 0.3*inch))
    
    # ========== EXECUTIVE SUMMARY ==========
    summary_heading = Paragraph("Executive Summary", heading_style)
    story.append(summary_heading)
    
    status = 'PASSED' if overall['errorRate'] < 1.0 else 'FAILED'
    status_color = colors.green if status == 'PASSED' else ERROR_RED
    
    summary_data = [
        ['Metric', 'Value', 'Status'],
        ['Error Rate', f"{overall['errorRate']:.2f}%", 
         'PASS' if overall['errorRate'] < 1.0 else 'FAIL'],
        ['Avg Response Time', f"{overall['avgResponseTime']:.2f} ms", 
         'PASS' if overall['avgResponseTime'] < 1000 else 'WARN'],
        ['Total Samples', str(overall['totalSamples']), 'INFO'],
        ['Total Errors', str(overall['totalErrors']), 
         'PASS' if overall['totalErrors'] == 0 else 'FAIL'],
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2.5*inch, 1.5*inch])
    
    # Build style list dynamically
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), SKY_BLUE),
        ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]
    
    # Add error highlighting for error row
    if overall['totalErrors'] > 0:
        style_commands.append(('BACKGROUND', (0, 4), (-1, 4), ERROR_RED))
        style_commands.append(('TEXTCOLOR', (0, 4), (-1, 4), WHITE))
        style_commands.append(('FONT', (0, 4), (-1, 4), 'Helvetica-Bold', 10))
    
    summary_table.setStyle(TableStyle(style_commands))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    
    # ========== RESPONSE TIME CHART ==========
    chart_heading = Paragraph("Response Time Analysis", heading_style)
    story.append(chart_heading)
    story.append(create_response_time_chart(transactions))
    story.append(Spacer(1, 0.3*inch))
    
    # ========== TRANSACTIONS PER SECOND CHART ==========
    tps_heading = Paragraph("Transactions Per Second (TPS)", heading_style)
    story.append(tps_heading)
    story.append(create_tps_chart(transactions))
    story.append(Spacer(1, 0.3*inch))
    
    # ========== ACTIVE THREADS OVER TIME CHART ==========
    threads_heading = Paragraph("Active Threads Over Time", heading_style)
    story.append(threads_heading)
    story.append(create_threads_over_time_chart(transactions))
    story.append(Spacer(1, 0.3*inch))
    
    # ========== DETAILED RESULTS TABLE ==========
    results_heading = Paragraph("Detailed Transaction Results", heading_style)
    story.append(results_heading)
    
    results_data = [
        ['Transaction', 'Samples', 'Errors', 'Error%', 'Avg(s)', 
         'Min(s)', 'Max(s)', '90%(s)']
    ]
    
    for tx in transactions:
        # Use Paragraph for transaction name to enable word wrapping
        tx_name_para = Paragraph(tx['name'], ParagraphStyle(
            'TxName',
            parent=styles['Normal'],
            fontSize=8,
            fontName='Helvetica-Bold',
            textColor=DARK_TEXT,
            wordWrap='CJK'
        ))
        
        results_data.append([
            tx_name_para,
            str(tx['samples']),
            str(tx['errors']),
            f"{tx['errorPct']:.2f}",
            f"{tx['avgTime']/1000:.3f}",  # Convert to seconds
            f"{tx['minTime']/1000:.3f}",  # Convert to seconds
            f"{tx['maxTime']/1000:.3f}",  # Convert to seconds
            f"{tx['p90']/1000:.3f}"       # Convert to seconds
        ])
    
    results_table = Table(results_data, colWidths=[2.5*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch])
    
    # Base style
    style_commands = [
        # Transaction column header - Dark blue to make it more prominent
        ('BACKGROUND', (0, 0), (0, 0), DARK_BLUE),
        ('TEXTCOLOR', (0, 0), (0, 0), WHITE),
        ('FONT', (0, 0), (0, 0), 'Helvetica-Bold', 10),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        
        # Other column headers - Sky blue
        ('BACKGROUND', (1, 0), (-1, 0), SKY_BLUE),
        ('TEXTCOLOR', (1, 0), (-1, 0), WHITE),
        ('FONT', (1, 0), (-1, 0), 'Helvetica-Bold', 9),
        ('ALIGN', (1, 0), (-1, 0), 'CENTER'),
        
        # Transaction names in data rows - Bold
        ('FONT', (0, 1), (0, -1), 'Helvetica-Bold', 8),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('TEXTCOLOR', (0, 1), (0, -1), DARK_TEXT),
        
        # Other data columns
        ('FONT', (1, 1), (-1, -1), 'Helvetica', 8),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('TEXTCOLOR', (1, 1), (-1, -1), DARK_TEXT),
        
        # Alternating light blue rows
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT_SKY_BLUE]),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]
    
    # Highlight ONLY error cells (columns 2 and 3: Errors and Error%) with red background and white bold text
    for i, tx in enumerate(transactions, start=1):
        if tx['errors'] > 0:
            # Red background ONLY for error-related cells (columns 2 and 3)
            # Column 2 = Errors, Column 3 = Error%
            style_commands.append(('BACKGROUND', (2, i), (3, i), ERROR_RED))
            style_commands.append(('TEXTCOLOR', (2, i), (3, i), WHITE))
            style_commands.append(('FONT', (2, i), (3, i), 'Helvetica-Bold', 8))
    
    results_table.setStyle(TableStyle(style_commands))
    
    story.append(results_table)
    story.append(PageBreak())
    
    # ========== PERFORMANCE ANALYSIS ==========
    analysis_heading = Paragraph("Performance Analysis", heading_style)
    story.append(analysis_heading)
    
    # Best and worst performers
    sorted_by_avg = sorted(transactions, key=lambda x: x['avgTime'])
    best_performers = sorted_by_avg[:3]
    worst_performers = sorted_by_avg[-3:]
    
    analysis_data = [
        ['Category', 'Transaction', 'Avg Response Time (ms)']
    ]
    
    for tx in best_performers:
        analysis_data.append(['Best Performer', tx['name'], f"{tx['avgTime']:.2f}"])
    
    for tx in worst_performers:
        analysis_data.append(['Worst Performer', tx['name'], f"{tx['avgTime']:.2f}"])
    
    analysis_table = Table(analysis_data, colWidths=[1.5*inch, 3*inch, 2*inch])
    analysis_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), SKY_BLUE),
        ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 9),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT_SKY_BLUE]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(analysis_table)
    story.append(Spacer(1, 0.3*inch))
    
    # ========== ERROR ANALYSIS ==========
    if overall['totalErrors'] > 0:
        error_heading = Paragraph("Error Analysis", heading_style)
        story.append(error_heading)
        
        error_data = [['Transaction', 'Error Count', 'Error %', 'Error Codes']]
        
        for tx in transactions:
            if tx['errors'] > 0:
                # Format error codes
                error_codes_str = ', '.join([f"{code}({count})" for code, count in tx.get('error_codes', {}).items()])
                if not error_codes_str:
                    error_codes_str = 'N/A'
                
                # Use Paragraph for transaction name to enable word wrapping
                tx_name_para = Paragraph(tx['name'], ParagraphStyle(
                    'ErrorTxName',
                    parent=styles['Normal'],
                    fontSize=9,
                    fontName='Helvetica-Bold',
                    textColor=DARK_TEXT,
                    wordWrap='CJK'
                ))
                
                error_data.append([
                    tx_name_para,
                    str(tx['errors']),
                    f"{tx['errorPct']:.2f}%",
                    error_codes_str
                ])
        
        if len(error_data) > 1:
            error_table = Table(error_data, colWidths=[2*inch, 1*inch, 1*inch, 2.5*inch])
            
            error_style = [
                ('BACKGROUND', (0, 0), (-1, 0), SKY_BLUE),
                ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
                ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONT', (0, 1), (-1, -1), 'Helvetica', 9),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT_SKY_BLUE]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]
            
            # Highlight ONLY error cells (columns 1 and 2: Error Count and Error %)
            for i in range(1, len(error_data)):
                error_style.append(('BACKGROUND', (1, i), (2, i), ERROR_RED))
                error_style.append(('TEXTCOLOR', (1, i), (2, i), WHITE))
                error_style.append(('FONT', (1, i), (2, i), 'Helvetica-Bold', 9))
            
            error_table.setStyle(TableStyle(error_style))
            story.append(error_table)
            story.append(Spacer(1, 0.3*inch))
    
    # ========== RECOMMENDATIONS ==========
    rec_heading = Paragraph("Recommendations", heading_style)
    story.append(rec_heading)
    
    recommendations = []
    
    if overall['errorRate'] > 1.0:
        recommendations.append("• Investigate and fix errors - error rate exceeds 1%")
    
    if overall['avgResponseTime'] > 1000:
        recommendations.append("• Optimize response times - average exceeds 1 second")
    
    for tx in transactions:
        if tx['p95'] > 2000:
            recommendations.append(f"• Optimize '{tx['name']}' - 95th percentile exceeds 2 seconds")
    
    if not recommendations:
        recommendations.append("• All metrics within acceptable ranges")
        recommendations.append("• Continue monitoring performance trends")
    
    for rec in recommendations:
        p = Paragraph(rec, styles['Normal'])
        story.append(p)
        story.append(Spacer(1, 0.1*inch))
    
    # Build PDF
    doc.build(story)

# ============================================================================
# Excel Report Generation with Enhanced Styling
# ============================================================================

def generate_excel_report(transactions, overall, output_path, project_name, environment, tester_name, prepared_by="", company_logo=""):
    """Generate enhanced Excel report with charts and styling"""
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Summary Sheet
    ws_summary = wb.create_sheet("Summary", 0)
    create_excel_summary_sheet(ws_summary, transactions, overall, project_name, environment, tester_name, prepared_by, company_logo)
    
    # Create Detailed Results Sheet
    ws_details = wb.create_sheet("Detailed Results", 1)
    create_excel_details_sheet(ws_details, transactions)
    
    # Create Charts Sheet
    ws_charts = wb.create_sheet("Charts & Analysis", 2)
    create_excel_charts_sheet(ws_charts, transactions, overall)
    
    # Save workbook
    wb.save(output_path)

def create_excel_summary_sheet(ws, transactions, overall, project_name, environment, tester_name, prepared_by="", company_logo=""):
    """Create summary sheet with sky blue theme"""
    
    # Title
    ws['A1'] = 'Performance Test Report'
    ws['A1'].font = Font(size=18, bold=True, color=EXCEL_DARK_BLUE)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30
    
    # Sky blue background for title
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}1'].fill = PatternFill(start_color=EXCEL_SKY_BLUE, end_color=EXCEL_SKY_BLUE, fill_type='solid')
    
    # Add company logo if provided
    row = 2
    if company_logo and os.path.exists(company_logo):
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(company_logo)
            img.width = 150
            img.height = 75
            ws.add_image(img, f'A{row}')
            row += 4  # Skip rows for logo
        except:
            pass  # Skip if logo can't be loaded
    
    # Project Information Section
    row += 1
    ws[f'A{row}'] = 'Project Information'
    ws[f'A{row}'].font = Font(size=12, bold=True, color=EXCEL_WHITE)
    ws[f'A{row}'].fill = PatternFill(start_color=EXCEL_DARK_BLUE, end_color=EXCEL_DARK_BLUE, fill_type='solid')
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    info_data = [
        ['Project Name:', project_name],
        ['Environment:', environment],
        ['Tested By:', tester_name],
        ['Report Prepared By:', prepared_by if prepared_by else tester_name],
        ['Test Start:', overall['testStart'].strftime('%Y-%m-%d %H:%M:%S')],
        ['Test End:', overall['testEnd'].strftime('%Y-%m-%d %H:%M:%S')],
        ['Duration:', format_duration(overall['duration'])]
    ]
    
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, color=EXCEL_WHITE)
        ws[f'A{row}'].fill = PatternFill(start_color=EXCEL_SKY_BLUE, end_color=EXCEL_SKY_BLUE, fill_type='solid')
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = PatternFill(start_color=EXCEL_LIGHT_BLUE, end_color=EXCEL_LIGHT_BLUE, fill_type='solid')
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'B{row}:F{row}')
        
        # Add borders
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    # Executive Summary Section
    row += 1
    ws[f'A{row}'] = 'Executive Summary'
    ws[f'A{row}'].font = Font(size=12, bold=True, color=EXCEL_WHITE)
    ws[f'A{row}'].fill = PatternFill(start_color=EXCEL_DARK_BLUE, end_color=EXCEL_DARK_BLUE, fill_type='solid')
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    summary_data = [
        ['Total Samples:', overall['totalSamples'], 'INFO'],
        ['Total Errors:', overall['totalErrors'], 'FAIL' if overall['totalErrors'] > 0 else 'PASS'],
        ['Error Rate:', f"{overall['errorRate']:.2f}%", 'FAIL' if overall['errorRate'] > 1.0 else 'PASS'],
        ['Avg Response Time:', f"{overall['avgResponseTime']:.2f} ms", 'WARN' if overall['avgResponseTime'] > 1000 else 'PASS']
    ]
    
    for label, value, status in summary_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, color=EXCEL_WHITE)
        ws[f'A{row}'].fill = PatternFill(start_color=EXCEL_SKY_BLUE, end_color=EXCEL_SKY_BLUE, fill_type='solid')
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws[f'B{row}'] = value
        ws[f'C{row}'] = status
        
        # Apply error highlighting for failed items
        if status == 'FAIL':
            ws[f'B{row}'].fill = PatternFill(start_color=EXCEL_ERROR_RED, end_color=EXCEL_ERROR_RED, fill_type='solid')
            ws[f'B{row}'].font = Font(bold=True, color=EXCEL_WHITE)
            ws[f'C{row}'].fill = PatternFill(start_color=EXCEL_ERROR_RED, end_color=EXCEL_ERROR_RED, fill_type='solid')
            ws[f'C{row}'].font = Font(bold=True, color=EXCEL_WHITE)
        else:
            ws[f'B{row}'].fill = PatternFill(start_color=EXCEL_LIGHT_BLUE, end_color=EXCEL_LIGHT_BLUE, fill_type='solid')
            ws[f'C{row}'].fill = PatternFill(start_color=EXCEL_LIGHT_BLUE, end_color=EXCEL_LIGHT_BLUE, fill_type='solid')
        
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add borders
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

def create_excel_details_sheet(ws, transactions):
    """Create detailed results sheet with sky blue theme and error highlighting"""
    
    # Title
    ws['A1'] = 'Detailed Transaction Results'
    ws['A1'].font = Font(size=14, bold=True, color=EXCEL_WHITE)
    ws['A1'].fill = PatternFill(start_color=EXCEL_DARK_BLUE, end_color=EXCEL_DARK_BLUE, fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:K1')
    ws.row_dimensions[1].height = 25
    
    # Headers
    headers = ['Transaction', 'Samples', 'Errors', 'Error %', 'Avg (s)', 
               'Min (s)', 'Max (s)', '90% (s)', 'TPS']
    
    row = 2
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        
        # Make Transaction column header more prominent with darker blue
        if col_idx == 1:  # Transaction column
            cell.font = Font(bold=True, color=EXCEL_WHITE, size=11)
            cell.fill = PatternFill(start_color=EXCEL_DARK_BLUE, end_color=EXCEL_DARK_BLUE, fill_type='solid')
        else:
            cell.font = Font(bold=True, color=EXCEL_WHITE)
            cell.fill = PatternFill(start_color=EXCEL_SKY_BLUE, end_color=EXCEL_SKY_BLUE, fill_type='solid')
        
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Data rows
    row = 3
    for tx in transactions:
        has_errors = tx['errors'] > 0
        
        # Alternating row background color
        bg_color = EXCEL_LIGHT_BLUE if row % 2 == 0 else EXCEL_WHITE
        
        # Transaction name (Column 1) - Make it bold and prominent with text wrapping
        cell = ws.cell(row=row, column=1)
        cell.value = tx['name']
        cell.font = Font(bold=True, size=10)  # Bold font for transaction names
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # Enable text wrapping
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        
        # Numeric data - Convert milliseconds to seconds
        data_values = [
            tx['samples'],      # Column 2
            tx['errors'],       # Column 3 - Error Count
            f"{tx['errorPct']:.2f}",  # Column 4 - Error %
            f"{tx['avgTime']/1000:.3f}",  # Column 5 - Avg in seconds
            f"{tx['minTime']/1000:.3f}",  # Column 6 - Min in seconds
            f"{tx['maxTime']/1000:.3f}",  # Column 7 - Max in seconds
            f"{tx['p90']/1000:.3f}",      # Column 8 - 90% in seconds
            f"{tx['throughput']:.2f}"     # Column 9 - TPS
        ]
        
        for col_idx, value in enumerate(data_values, start=2):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Apply red highlighting ONLY to error-related cells (columns 3 and 4)
            if has_errors and col_idx in [3, 4]:  # Errors and Error% columns
                cell.fill = PatternFill(start_color=EXCEL_ERROR_RED, end_color=EXCEL_ERROR_RED, fill_type='solid')
                cell.font = Font(bold=True, color=EXCEL_WHITE)
            else:
                # Normal alternating background
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        
        # Add borders to all cells
        for col_idx in range(1, 10):  # Updated to 10 columns
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    # Set column widths for perfect alignment
    ws.column_dimensions['A'].width = 40  # Increased width for transaction names
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    # Set row heights to accommodate wrapped text
    for row_idx in range(3, row):
        ws.row_dimensions[row_idx].height = 30  # Adequate height for wrapped text

def create_excel_charts_sheet(ws, transactions, overall):
    """Create charts and analysis sheet"""
    
    # Title
    ws['A1'] = 'Performance Analysis & Charts'
    ws['A1'].font = Font(size=14, bold=True, color=EXCEL_WHITE)
    ws['A1'].fill = PatternFill(start_color=EXCEL_DARK_BLUE, end_color=EXCEL_DARK_BLUE, fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 25
    
    # Response Time Chart Data
    row = 3
    ws[f'A{row}'] = 'Response Time Analysis'
    ws[f'A{row}'].font = Font(size=12, bold=True, color=EXCEL_DARK_BLUE)
    
    row += 1
    ws[f'A{row}'] = 'Transaction'
    ws[f'B{row}'] = 'Avg Response Time (ms)'
    
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = Font(bold=True, color=EXCEL_WHITE)
        ws[f'{col}{row}'].fill = PatternFill(start_color=EXCEL_SKY_BLUE, end_color=EXCEL_SKY_BLUE, fill_type='solid')
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    chart_start_row = row
    
    # Add top 10 transactions by response time
    sorted_tx = sorted(transactions, key=lambda x: x['avgTime'], reverse=True)[:10]
    for tx in sorted_tx:
        ws[f'A{row}'] = tx['name'][:30]  # Truncate long names
        ws[f'B{row}'] = tx['avgTime']
        ws[f'B{row}'].number_format = '0.00'
        row += 1
    
    chart_end_row = row - 1
    
    # Create Bar Chart for Response Times
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Average Response Time by Transaction"
    chart.y_axis.title = 'Response Time (ms)'
    chart.x_axis.title = 'Transaction'
    
    data = Reference(ws, min_col=2, min_row=chart_start_row-1, max_row=chart_end_row)
    cats = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    
    ws.add_chart(chart, f'D{chart_start_row}')
    
    # TPS Analysis Section
    row = chart_end_row + 3
    ws[f'A{row}'] = 'Transactions Per Second (TPS)'
    ws[f'A{row}'].font = Font(size=12, bold=True, color=EXCEL_DARK_BLUE)
    
    row += 1
    ws[f'A{row}'] = 'Transaction'
    ws[f'B{row}'] = 'TPS'
    
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = Font(bold=True, color=EXCEL_WHITE)
        ws[f'{col}{row}'].fill = PatternFill(start_color=EXCEL_SKY_BLUE, end_color=EXCEL_SKY_BLUE, fill_type='solid')
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    tps_start_row = row
    
    # Add top 10 transactions by TPS
    sorted_tps = sorted(transactions, key=lambda x: x['throughput'], reverse=True)[:10]
    for tx in sorted_tps:
        ws[f'A{row}'] = tx['name'][:30]
        ws[f'B{row}'] = tx['throughput']
        ws[f'B{row}'].number_format = '0.00'
        row += 1
    
    tps_end_row = row - 1
    
    # Error Analysis Section
    row = tps_end_row + 3
    ws[f'A{row}'] = 'Error Analysis'
    ws[f'A{row}'].font = Font(size=12, bold=True, color=EXCEL_DARK_BLUE)
    
    row += 1
    ws[f'A{row}'] = 'Transaction'
    ws[f'B{row}'] = 'Error Count'
    ws[f'C{row}'] = 'Error %'
    ws[f'D{row}'] = 'Error Codes'
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True, color=EXCEL_WHITE)
        ws[f'{col}{row}'].fill = PatternFill(start_color=EXCEL_SKY_BLUE, end_color=EXCEL_SKY_BLUE, fill_type='solid')
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    error_start_row = row
    
    # Add transactions with errors
    error_txs = [tx for tx in transactions if tx['errors'] > 0]
    if error_txs:
        for idx, tx in enumerate(error_txs):
            # Alternating background for transaction name
            bg_color = EXCEL_LIGHT_BLUE if idx % 2 == 0 else EXCEL_WHITE
            
            # Transaction name - normal alternating color with word wrap
            ws[f'A{row}'] = tx['name']
            ws[f'A{row}'].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Error Count - RED highlighting
            ws[f'B{row}'] = tx['errors']
            ws[f'B{row}'].fill = PatternFill(start_color=EXCEL_ERROR_RED, end_color=EXCEL_ERROR_RED, fill_type='solid')
            ws[f'B{row}'].font = Font(bold=True, color=EXCEL_WHITE)
            ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            # Error % - RED highlighting
            ws[f'C{row}'] = tx['errorPct']
            ws[f'C{row}'].number_format = '0.00'
            ws[f'C{row}'].fill = PatternFill(start_color=EXCEL_ERROR_RED, end_color=EXCEL_ERROR_RED, fill_type='solid')
            ws[f'C{row}'].font = Font(bold=True, color=EXCEL_WHITE)
            ws[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            # Error Codes
            error_codes_str = ', '.join([f"{code}({count})" for code, count in tx.get('error_codes', {}).items()])
            if not error_codes_str:
                error_codes_str = 'N/A'
            ws[f'D{row}'] = error_codes_str
            ws[f'D{row}'].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            ws[f'D{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Add borders
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            ws.row_dimensions[row].height = 30  # Height for wrapped text
            row += 1
    else:
        ws[f'A{row}'] = 'No errors detected'
        ws[f'A{row}'].font = Font(color='00FF00')
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30

# ============================================================================
# Main Execution
# ============================================================================


# ============================================================================
# Main Execution
# ============================================================================

def main():
    """Main execution function"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate enhanced JMeter performance test reports')
    parser.add_argument('input_file', help='Path to JMeter JTL/CSV results file')
    parser.add_argument('--project', default='Performance Test', help='Project name')
    parser.add_argument('--environment', default='Test Environment', help='Test environment')
    parser.add_argument('--tester', default='Performance Tester', help='Tester name')
    parser.add_argument('--prepared-by', default='', help='Report prepared by')
    parser.add_argument('--logo', default='', help='Company logo path')
    parser.add_argument('--output-dir', default='test_reports', help='Output directory for reports')
    
    args = parser.parse_args()
    
    print("=" * 80)
    print("📊 JMeter Enhanced Report Generator v2.0")
    print("=" * 80)
    print(f"Input file: {args.input_file}")
    print(f"Project: {args.project}")
    print(f"Environment: {args.environment}")
    print(f"Tester: {args.tester}")
    print()
    
    # Create output directory
    os.makedirs(args.output_dir, exist_ok=True)
    
    # Parse JMeter results
    print("📖 Parsing JMeter results...")
    parser = JMeterResultParser(args.input_file)
    transactions, overall = parser.parse()
    
    print(f"✅ Parsed {overall['totalSamples']} samples from {len(transactions)} transactions")
    print()
    
    # Generate timestamp for filenames
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    base_name = f"Performance_Test_Report_{timestamp}"
    
    # Generate PDF report
    pdf_path = os.path.join(args.output_dir, f"{base_name}.pdf")
    print(f"📄 Generating PDF report: {pdf_path}")
    
    try:
        generate_pdf_report(
            transactions, overall, pdf_path,
            args.project, args.environment, args.tester,
            args.prepared_by, args.logo
        )
        print("✅ PDF report generated successfully!")
    except Exception as e:
        print(f"❌ Error generating PDF: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print()
    
    # Generate Excel report
    excel_path = os.path.join(args.output_dir, f"{base_name}.xlsx")
    print(f"📊 Generating Excel report: {excel_path}")
    
    try:
        generate_excel_report(
            transactions, overall, excel_path,
            args.project, args.environment, args.tester,
            args.prepared_by, args.logo
        )
        print("✅ Excel report generated successfully!")
    except Exception as e:
        print(f"❌ Error generating Excel: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print()
    print("🎉 Report generation complete!")
    print(f"📁 Reports saved to: {args.output_dir}")
    print("=" * 80)
    print()
    print("Features included:")
    print("  🎨 Sky blue professional theme")
    print("  🔴 Error highlighting (red background + white bold text)")
    print("  📊 Interactive charts in Excel")
    print("  📐 Perfect alignment and borders")
    print("  📊 Multi-sheet Excel workbook")
    print()
    print("Open the reports to see the enhanced features!")
    print("=" * 80)


if __name__ == '__main__':
    main()
